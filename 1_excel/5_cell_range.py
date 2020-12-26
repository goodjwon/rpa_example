from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# if english teacher
col_B = ws["B"]
#print(col_B)

print("==========get BCol===============")
for cell in col_B:
    print(cell.value)

print("==========get Cols===============")
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value)

print("==========get Title===============")
row_title = ws[1]
for cell in row_title:
    print(cell.value)

print("==========get Rows================")
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print("==========get MaxRows==============")
row_range = ws[2: ws.max_row]
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        print(xy)
    print()

print("==========get AllRows==============")
print(tuple(ws.rows))

for row in tuple(ws.rows):
    print(row[1].value)

for row in ws.iter_rows():
    print(row[1].value)


print("==========get AllColumns==============")
print(tuple(ws.columns))

for column in tuple(ws.columns):
    print(column[0].value)

for column in ws.iter_cols():
    print(column[0].value)

print("==========get 1row to 5 and 2 col to 3 =====")
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어



wb.save("sample4.xlsx")
